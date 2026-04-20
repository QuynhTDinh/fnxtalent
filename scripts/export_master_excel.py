import os
import sys
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Protection, Border, Side
from openpyxl.utils import get_column_letter

project_root = Path(__file__).parent.parent
sys.path.append(str(project_root))

from core.taxonomy.loader import get_taxonomy

def export_to_excel():
    print("Loading FNX Taxonomy...")
    tax = get_taxonomy()
    
    comp_to_tld_vi = {}
    for zone_id, z in tax.tld_zones.items():
        for cid in z.competency_ids:
            comp_to_tld_vi[cid] = z.name 
            
    # 1. CORE LIBRARY
    core_data = []
    sector = "CORE"
    
    comp_cols = {} # Mapping K1 -> "K1 (Kiến thức chuyên ngành)" for Matrix
    for grp in tax.ask_groups:
        for comp in grp.competencies:
            comp_cols[comp.id] = f"{comp.id} ({comp.name})"
            tld_name = comp_to_tld_vi.get(comp.id, "N/A")
            ask_name = f"{grp.id} - {grp.vi}"
            
            for lv_num, lv_desc in comp.levels.items():
                # Lấy trực tiếp định danh cấp bậc (1-6) thay vì map string bị hụt số 2
                bloom_level = int(lv_num)

                core_data.append({
                    "Unique ID": f"FNX-{sector}-{comp.katz_zone[0] if comp.katz_zone else 'G'}-{comp.id}-{lv_num}",
                    "Layer 1 (A-S-K)": ask_name,
                    "Layer 2 (T-L-D)": tld_name,
                    "Competency Name": comp.name,
                    "Bloom Level": bloom_level,
                    "Action Descriptor": tax.levels[lv_num].vi if lv_num in tax.levels else "",
                    "Criticality (1-5)": 3, 
                    "Evidence Indicators (Điền biểu hiện Tích cực)": lv_desc,
                    "Negative Indicators [Red Flags] (Điền biểu hiện Tiêu cực)": "TBD - Cần AI/HR bổ sung"
                })
    df_core = pd.DataFrame(core_data)
    
    # 2. ROLE BENCHMARKS
    role_data = []
    for role_id, profile in tax.role_profiles.items():
        role_data.append({
            "Role ID": role_id,
            "Department": profile.department_name,
            "Job Title/Role Name": profile.name,
            "Expected Level": profile.expected_level,
            "Target Knowledge (K%)": profile.ask_weights.get("K", 0),
            "Target Skill (S%)": profile.ask_weights.get("S", 0),
            "Target Attitude (A%)": profile.ask_weights.get("A", 0),
            "Layer 2: Technique (T%)": profile.katz_weights.get("TECHNICAL", 0),
            "Layer 2: Language (L%)": profile.katz_weights.get("INTERPERSONAL", 0),
            "Layer 2: Digital (D%)": profile.katz_weights.get("CONCEPTUAL", 0)
        })
    df_roles = pd.DataFrame(role_data)

    # 3. COMPETENCY MATRIX (The New Layer 3 Sheet)
    matrix_data = []
    for role_id, profile in tax.role_profiles.items():
        row = {
            "Department": profile.department_name,
            "Job Title": profile.job_title_name,
            "Career Level": profile.name,
            "Expected Band": profile.expected_level
        }
        # Thêm 11 cột chấm điểm năng lực
        for cid, cname in comp_cols.items():
            row[cname] = profile.competency_targets.get(cid, 0)
        matrix_data.append(row)
    df_matrix = pd.DataFrame(matrix_data)
    
    # 4. COMBAT SCENARIOS
    combat_data = [{
        "Combat ID": f"FNX-{sector}-COMB-001",
        "Target Role ID": "pmb_manager_pro",
        "Scenario Name": "Xử lý sự cố áp suất tháp",
        "Context/Background": "Hệ thống báo động đỏ, áp suất tăng đột biến trong ca đêm.",
        "Challenge (Trigger)": "Van xả kép bị kẹt, hệ thống dự phòng không kích hoạt.",
        "Key Competencies Assessed": "K2, S2, A1",
        "Expected Behaviors": "Ngừng hệ thống an toàn, báo cáo cấp trên.",
        "Red Flags (Automatic Fail)": "Bỏ qua cảnh báo, hoảng loạn."
    }]
    df_combat = pd.DataFrame(combat_data)
    
    # EXPORT VÀ APPLY STYLING
    export_dir = project_root / "data" / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    output_path = export_dir / "SSOT_Master_Taxonomy.xlsx"
    
    print(f"Exporting and formatting to {output_path}...")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_core.to_excel(writer, sheet_name="Core_Library", index=False)
        df_roles.to_excel(writer, sheet_name="Role_Benchmarks", index=False)
        df_matrix.to_excel(writer, sheet_name="Competency_Matrix", index=False)
        df_combat.to_excel(writer, sheet_name="Combat_Scenarios", index=False)
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        input_target_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Màu vàng chanh cho Matrix
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Format Core Library
        worksheet = writer.sheets['Core_Library']
        worksheet.freeze_panes = 'A2'
        widths = [18, 15, 20, 25, 12, 35, 15, 50, 50]
        for col_idx, width in enumerate(widths, 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = width
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        current_comp = ""
        merge_start_row = 2
        for row in range(2, len(df_core) + 2):
            for col in range(1, 10):
                c = worksheet.cell(row=row, column=col)
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.border = thin_border
            comp_name = worksheet.cell(row=row, column=4).value
            if comp_name != current_comp:
                if current_comp != "":
                    for merge_col in [2, 3, 4]:
                        worksheet.merge_cells(start_row=merge_start_row, start_column=merge_col, end_row=row-1, end_column=merge_col)
                        m_cell = worksheet.cell(row=merge_start_row, column=merge_col)
                        m_cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                        if merge_col == 4: m_cell.font = Font(bold=True)
                current_comp = comp_name
                merge_start_row = row
        for merge_col in [2, 3, 4]:
            worksheet.merge_cells(start_row=merge_start_row, start_column=merge_col, end_row=len(df_core)+1, end_column=merge_col)
            m_cell = worksheet.cell(row=merge_start_row, column=merge_col)
            m_cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            if merge_col == 4: m_cell.font = Font(bold=True)
        worksheet.column_dimensions['A'].hidden = True
        
        # Format Role Benchmarks
        worksheet2 = writer.sheets['Role_Benchmarks']
        worksheet2.freeze_panes = 'A2'
        widths2 = [10, 20, 25, 12, 12, 12, 12, 15, 15, 15]
        for col_idx, width in enumerate(widths2, 1):
            col_letter = get_column_letter(col_idx)
            worksheet2.column_dimensions[col_letter].width = width
            cell = worksheet2.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Format Competency Matrix
        worksheet3 = writer.sheets['Competency_Matrix']
        worksheet3.freeze_panes = 'E2' # Khóa 4 cột đầu
        worksheet3.protection.sheet = True # Bật khóa

        widths3 = [20, 25, 20, 12] + [12]*11
        for col_idx, width in enumerate(widths3, 1):
            col_letter = get_column_letter(col_idx)
            worksheet3.column_dimensions[col_letter].width = width
            cell = worksheet3.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in range(2, len(df_matrix) + 2):
            for col in range(1, len(widths3) + 1):
                c = worksheet3.cell(row=row, column=col)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border
                c.protection = Protection(locked=True)
                
                # Mở khóa 11 cột Năng lực và tô màu Vàng nê-ông báo hiệu
                if col >= 5:
                    c.protection = Protection(locked=False)
                    c.fill = input_target_fill

    print("✅ Formatted Export completed successfully!")

if __name__ == "__main__":
    export_to_excel()
